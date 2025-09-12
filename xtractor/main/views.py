import re
import os
import json, tempfile, os
import pprint
import requests
import numpy as np
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from django.urls import reverse
from django.conf import settings
from django.db import transaction
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from django.contrib.auth.views import LoginView
from django.contrib.auth.hashers import check_password
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout, update_session_auth_hash
from django.http import JsonResponse, HttpResponseRedirect, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from .services.azure_vision import analyze_image
from .models import  ComponentType, FormObject, HeaderObjects, RowObjects, FieldObject, FormName, Extraction, ExtractedFields, ExtractedTable
from .forms import TypeForm, FormObjectForm, HeaderObjectsForm, RowObjectsForm, FieldObjectForm, ChangePasswordForm



class CustomLoginView(LoginView):
    template_name = "login.html"

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            return redirect("main:index")  # <-- change "index" to your URL name
        return super().dispatch(request, *args, **kwargs)


@login_required
def change_password(request):
    if request.method == "POST":
        form = ChangePasswordForm(request.POST)
        if form.is_valid():
            old_password = form.cleaned_data.get("old_password")
            new_password1 = form.cleaned_data.get("new_password1")
            new_password2 = form.cleaned_data.get("new_password2")

            if not check_password(old_password, request.user.password):
                messages.error(request, "Old password is incorrect.")
            elif new_password1 != new_password2:
                messages.error(request, "New passwords do not match.")
            elif len(new_password1) < 8:
                messages.error(request, "New password must be at least 8 characters.")
            else:
                request.user.set_password(new_password1)
                request.user.save()
                update_session_auth_hash(request, request.user)  # ✅ keeps user logged in
                messages.success(request, "Password changed successfully!")
                return redirect("main:index")  # change "index" to your homepage name
    else:
        form = ChangePasswordForm()

    return render(request, "change_password.html", {"form": form})



def login(request):
    
    return render(request, 'login.html')


def logout_request(request):
    logout(request)   # flushes the session
    request.session.flush()  # just to be sure
    return redirect("main:login_request")

@login_required
def index(request):
    return render(request, 'index.html')

    
    
@login_required   
def template_list(request):
    
    return render(request, 'template_list.html')


@csrf_exempt
@login_required
def load_template_list(request):
    draw = int(request.POST.get('draw', 1))
    start = int(request.POST.get('start', 0))
    length = int(request.POST.get('length', 10))
    search_value = request.POST.get('search[value]', '')
    order_column_index = request.POST.get('order[0][column]')
    order_direction = request.POST.get('order[0][dir]', 'asc')

    # 🔹 Query all records from DB
    qs = FormName.objects.filter(status=1)

    # 🔹 Filtering (LIKE query)
    if search_value:
        qs = qs.filter(name__icontains=search_value)

    total_records = FormName.objects.count()   # total in DB
    filtered_records = qs.count()              # total after filtering

    # 🔹 Sorting
    if order_column_index is not None:
        order_column_index = int(order_column_index)
        order_column_name = request.POST.get(f'columns[{order_column_index}][data]')
        if order_column_name:
            if order_direction == "desc":
                qs = qs.order_by(f'-{order_column_name}')
            else:
                qs = qs.order_by(order_column_name)

    # 🔹 Pagination
    qs = qs[start:start + length]

    # 🔹 Convert queryset to JSON
    data = list(qs.values("id", "name"))

    response = {
        "draw": draw,
        "recordsTotal": total_records,
        "recordsFiltered": filtered_records,
        "data": data,
    }
    return JsonResponse(response)

    
@login_required   
def template_config(request, pk=None):
    
    return render(request, "config.html")



def submit_form_ajax(request):
    if request.method != 'POST':
        return JsonResponse({"success": False, "error": "Invalid request method"}, status=405)

    try:
        with transaction.atomic():
            form_fields = request.POST.getlist('field_list[]')
            form_data = dict(request.POST)


            #check if there is at least field or table
            if not form_fields and not any(k.startswith("table_form") for k in form_data.keys()):
                return JsonResponse({"success": False, "error": "Please add table or field in the form"}, status=400)
            
            #get form title
            form_title = request.POST.get('form_title')
            if not form_title:
                return JsonResponse({"success": False, "error": "form_title is required"}, status=400)

            #check if already exist or create
            form_name, created = FormName.objects.get_or_create(name=form_title)
            if not created:
                return JsonResponse({"success": False, "error": "Form already exists"}, status=400)
            
            #saving of fields    
            if form_fields:
                component_field = ComponentType.objects.get(type="Field")
                for ff in form_fields:
                    t1 = FormObject(form_name=form_name, type=component_field, title=ff)
                    t1.save()

            #saving of tables
            tables = defaultdict(dict)

            #fixing tables
            pattern = re.compile(r"table_form\[(\d+)\]\[(.+?)\](?:\[\])?$")

            for key, value in form_data.items():
                match = pattern.match(key)
                if match:
                    index, field = match.groups()

                    if "header][label]" in key and not key.endswith("[]"):
                        tables[index]["label_header"] = value[0]

                    elif "header][label" in key and key.endswith("[]"):
                        tables[index].setdefault("labels", []).extend(value)

                    elif field == "header" and key.endswith("[]"):
                        tables[index].setdefault("headers", []).extend(value)

                    else:
                        tables[index][field] = value[0] if len(value) == 1 else value

            tables = dict(tables)
            
            for k, v in tables.items():
                component_table = ComponentType.objects.get(type="Table")

                t1 = FormObject(form_name=form_name, type=component_table, title=v['table_title'])
                t1.save()

                if v['label_header'].strip() != '':
                    header_label = HeaderObjects(form_object=t1, header_name=v['label_header'], header_type='label')
                    header_label.save()

                else:
                    header_label = HeaderObjects(form_object=t1, header_name="NA", header_type='label')
                    header_label.save()

                for l in v['labels']:
                    row=RowObjects(form_object=t1, row_name=l)
                    row.save()
                for h in v['headers']:
                    header = HeaderObjects(form_object=t1, header_name=h, header_type='value')
                    header.save()

         
            return JsonResponse({
                "success": True,
                "redirect_url": reverse("main:template_detail", args=[form_name.pk])
                
            })

    except Exception as e:
        import traceback
        traceback.print_exc()  # prints full stack trace in console
        return JsonResponse({"success": False, "error": str(e)}, status=500)



def edit_form_ajax(request, pk):
    if request.method != 'POST':
        return JsonResponse({"success": False, "error": "Invalid request method"}, status=405)

    try:
        

        with transaction.atomic():
            form_fields = request.POST.getlist('field_list[]')
          
            
            form_data = dict(request.POST)

            # check if there is at least one field or table
            if not form_fields and not any(k.startswith("table_form") for k in form_data.keys()):
                return JsonResponse({"success": False, "error": "Please add table or field in the form"}, status=400)

            # fetch the existing form
            try:
                form_name = FormName.objects.get(pk=pk)
            except FormName.DoesNotExist:
                return JsonResponse({"success": False, "error": "Form does not exist"}, status=404)

            # update form title if provided
            form_title = request.POST.get('form_title')
            if form_title:
                form_name.name = form_title
                form_name.save()

            # delete old FormObjects (and cascade their headers/rows)
            FormObject.objects.filter(form_name=form_name).delete()

            # re-save fields
            if form_fields:
                component_field = ComponentType.objects.get(type="Field")
                for ff in form_fields:
                    t1 = FormObject(form_name=form_name, type=component_field, title=ff)
                    t1.save()

            # re-save tables
            tables = defaultdict(dict)
            pattern = re.compile(r"table_form\[(\d+)\]\[(.+?)\](?:\[\])?$")

            for key, value in form_data.items():
                match = pattern.match(key)
                if match:
                    index, field = match.groups()

                    if "header][label]" in key and not key.endswith("[]"):
                        tables[index]["label_header"] = value[0]

                    elif "header][label" in key and key.endswith("[]"):
                        tables[index].setdefault("labels", []).extend(value)

                    elif field == "header" and key.endswith("[]"):
                        tables[index].setdefault("headers", []).extend(value)

                    else:
                        tables[index][field] = value[0] if len(value) == 1 else value

            tables = dict(tables)

            for k, v in tables.items():
                component_table = ComponentType.objects.get(type="Table")
                t1 = FormObject(form_name=form_name, type=component_table, title=v['table_title'])
                t1.save()

                if v['label_header'].strip() != '':
                    header_label = HeaderObjects(form_object=t1, header_name=v['label_header'], header_type='label')
                    header_label.save()
                else:
                    header_label = HeaderObjects(form_object=t1, header_name="NA", header_type='label')
                    header_label.save()

                for l in v['labels']:
                    row = RowObjects(form_object=t1, row_name=l)
                    row.save()

                for h in v['headers']:
                    header = HeaderObjects(form_object=t1, header_name=h, header_type='value')
                    header.save()

            return JsonResponse({
                "success": True,
                "redirect_url": reverse("main:template_detail", args=[form_name.pk])
            })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({"success": False, "error": str(e)}, status=500)



@csrf_exempt
def disable_form_ajax(request, pk):
    if request.method == "POST":  # ensure it's called via POST
        try:
            form = FormName.objects.get(pk=pk)
            form.status = 0  # disable
            form.save()
            return JsonResponse({"success": True, "message": f"{form.name} disabled successfully."})
        except FormName.DoesNotExist:
            raise Http404("Form not found")
    return JsonResponse({"success": False, "message": "Invalid request method"})


@login_required
def extractor(request):
    forms = FormName.objects.filter(status=1)
   
    return render(request, 'extractor.html', {"forms": forms})


@login_required
def submit_form_extractor(request):
    
    return render(request, 'extractor.html')



def template_detail(request, pk):
    form = get_object_or_404(
        FormName.objects.prefetch_related(
            'formobject_set__type',
            'formobject_set__headers',
            'formobject_set__rows',
        ),
        pk=pk
    )


    context = {
        "form": form,
        "form_objects": form.formobject_set.all()
    }
    
    
    return render(request, "components.html", context)




def compute_ocr_reliability(ocr_json, low_conf_threshold=0.5):
    """
    Compute a reliability score for Azure OCR-style JSON with explainability breakdown.
    """
    width = ocr_json.get("metadata", {}).get("width", 1)
    height = ocr_json.get("metadata", {}).get("height", 1)
    img_area = width * height

    words = []
    for block in ocr_json.get("readResult", {}).get("blocks", []):
        for line in block.get("lines", []):
            for w in line.get("words", []):
                words.append(w)

    if not words:
        return {
            "score": 0,
            "breakdown": {
                "average_conf": 0,
                "low_conf_ratio": 1,
                "density": 0,
                "num_words": 0,
                "confidence_contrib": 0,
                "low_conf_contrib": 0,
                "density_contrib": 0
            }
        }

    # Confidence values
    confs = [w["confidence"] for w in words if "confidence" in w]
    avg_conf = float(np.mean(confs))

    # Low-confidence ratio
    low_conf_ratio = float(np.mean([c < low_conf_threshold for c in confs])) if confs else 1

    # Text density (characters per pixel)
    total_chars = sum(len(w.get("text", "")) for w in words)
    density = total_chars / img_area if img_area > 0 else 0

    # Contributions
    conf_contrib = avg_conf * 70
    low_conf_contrib = (1 - low_conf_ratio) * 20
    density_contrib = min(density * 1e6, 10)

    score = conf_contrib + low_conf_contrib + density_contrib

    return {
        "score": round(min(score, 100), 2),
        "breakdown": {
            "average_conf": round(avg_conf, 3),
            "low_conf_ratio": round(low_conf_ratio, 3),
            "density": round(density, 8),
            "num_words": len(words),
            "confidence_contrib": round(conf_contrib, 2),
            "low_conf_contrib": round(low_conf_contrib, 2),
            "density_contrib": round(density_contrib, 2)
        }
    }


@csrf_exempt  # remove if you’re handling CSRF properly
def get_ave(request):
    """
    Handles uploaded image file and returns OCR reliability score with breakdown.
    Accepts multipart form-data with 'file' field.
    """
    image_file = request.FILES.get("file")
    if not image_file:
        return JsonResponse({"error": "Missing uploaded image file"}, status=400)

    # Call Azure OCR
    ocr_json = analyze_image(image_file)

    # If Azure OCR returned an error
    if "error" in ocr_json:
        return JsonResponse({"error": ocr_json["error"]}, status=500)

    # Compute reliability score
    result = compute_ocr_reliability(ocr_json)
    
    return JsonResponse(result, json_dumps_params={"indent": 2})


def upload_form(request):
    return render(request, 'ocr_upload.html')











@login_required   
def template_config(request, pk=None):
    
    return render(request, "config.html")





def get_values(counter, key, value):

    return 


























def sample(request):
    
    return render(request, 'sample.html')



# def template_config(request):
#     # if this is a POST request we need to process the form data
#     if request.method == "POST":
#         # create a form instance and populate it with data from the request:
#         typeform = TypeForm(request.POST, prefix='type')
#         formobjectform = FormObjectForm(request.POST, prefix='formobject')
#         headerobjectsform = HeaderObjectsForm(request.POST, prefix='headerobjects')
#         rowobjectsform = RowObjectsForm(request.POST, prefix='rowobjects')
#         fieldobjectform = FieldObjectForm(request.POST, prefix='fieldobject')
        
#         if 'submit_TypeForm' in request.POST:
#             if(typeform.is_valid()):
#                 #process type form
#                 type = typeform.save(commit=False) # Don't save yet
#                 typeform.save()  # Save type                
#         elif 'submit_FormObject' in request.POST:
#             if(formobjectform.is_valid()):
#                 #process form object
#                 formobject = formobjectform.save()  # Save form   
#         elif 'submit_HeaderObjects' in request.POST:    
#             if(headerobjectsform.is_valid()):
#                 #process form object 
#                 headerobjects = headerobjectsform.save()  # Save header   
#         elif 'submit_RowObjects' in request.POST:    
#             if(rowobjectsform.is_valid()):   
#                 #process form object
#                 rowobjects = rowobjectsform.save()  # Save row  
#         elif 'submit_FieldObject' in request.POST:    
#             if(fieldobjectform.is_valid()):   
#                 #process form object 
#                 fieldobject = fieldobjectform.save()  # Save row  
#     else :
#         typeform = TypeForm(prefix='type')
#         formobjectform = FormObjectForm(prefix='formobject')
#         headerobjectsform = HeaderObjectsForm(prefix='headerobjects')
#         rowobjectsform = RowObjectsForm(prefix='rowobjects')
#         fieldobjectform = FieldObjectForm(prefix='fieldobject')
   
#     return render(request, 'config.html', {'typeform': typeform, 'formobjectform': formobjectform,  'headerobjectsform': headerobjectsform,  'rowobjectsform': rowobjectsform,  'fieldobjectform': fieldobjectform})


   


def is_same_column(header_bbox, cell_bbox):
    header_xs = [p["x"] for p in header_bbox]
    cell_xs = [p["x"] for p in cell_bbox]

    header_left = min(header_xs)
    header_right = max(header_xs)

    cell_left = min(cell_xs)
    cell_right = max(cell_xs)

    # Overlap condition (any horizontal overlap)
    overlaps = not (cell_right < header_left or cell_left > header_right)

    # Fully contained condition (cell box entirely inside header box)
    fully_within = (cell_left >= header_left and cell_right <= header_right)

    # Return True if either overlaps or fully contained
    return overlaps or fully_within



def get_center(bbox):
    xs = [p["x"] for p in bbox]
    ys = [p["y"] for p in bbox]
    return (sum(xs) / 4, sum(ys) / 4)






@csrf_exempt
@login_required
def ocr_result_view(request):
    if request.method != "POST" or "file" not in request.FILES:
        return JsonResponse({"error": "No file uploaded"}, status=400)

    file = request.FILES["file"]
    template_id = request.POST.get("template_id")
    if not template_id:
        return JsonResponse({"error": "No template selected"}, status=400)

    ocr_data = analyze_image(file)  # call_azure_ocr reads the file bytes
    
    if "error" in ocr_data:
        return JsonResponse({"error": ocr_data["error"]}, status=400)

    form_name = FormName.objects.get(pk=template_id)

    lines = []  # line of texts extracted
    for block in ocr_data["readResult"]["blocks"]:
        for line in block["lines"]:
            center = get_center(line["boundingPolygon"])
            # Collect word-level confidences
            word_confs = [w.get("confidence", 1.0) for w in line.get("words", [])]
            avg_conf = np.mean(word_confs) if word_confs else 1.0

            lines.append({
                "text": line["text"],
                "center": center,
                "bbox": line["boundingPolygon"],
                "confidence": round(float(avg_conf), 3)  # keep 3 decimals
            })

    # === EXTRA TOTALS ===
    fields_to_find = list(
        FormObject.objects
        .filter(form_name=form_name, type__type="Field")
        .values_list('title', flat=True)
    )

    extra_totals = {}

    for field in fields_to_find:
        for line in lines:
            text_upper = line["text"]
            if text_upper.startswith(field):
                rest = text_upper[len(field):].strip()
                if rest:
                    extra_totals[field] = {
                        "text": rest,
                        "confidence": line.get("confidence", 1.0)
                    }
                else:
                    label_x, label_y = line["center"]
                    candidates = [
                        c for c in lines
                        if c != line and c["center"][0] > label_x and abs(c["center"][1] - label_y) < 10
                    ]
                    if candidates:
                        candidates.sort(key=lambda c: c["center"][0])
                        chosen = candidates[0]
                        extra_totals[field] = {
                            "text": chosen["text"],
                            "confidence": chosen.get("confidence", 1.0)
                        }
                    else:
                        extra_totals[field] = {"text": "N/A", "confidence": 0.0}
                break
    
    # === TABLES ===
    tables = []
    table_name = list(
        FormObject.objects
        .filter(form_name=form_name, type__type="Table")
        .values_list('title', flat=True)
    )
  
    for t in table_name:
        table = FormObject.objects.get(form_name=form_name,title=t)
        reject_anchor_y = None

        for line in lines:
            if line["text"] == table.title:
                reject_anchor_x, reject_anchor_y = line["center"]
                break

        label_header = (
            HeaderObjects.objects
            .filter(form_object=table, header_type="label")
            .values_list("header_name", flat=True)
            .first()
        ) or "LABEL"

        header_names = list(
            HeaderObjects.objects
            .filter(form_object=table)
            .exclude(header_type="label")
            .values_list('header_name', flat=True)
        )

        header_x_positions = {}
        if reject_anchor_y is not None:
            closest_headers = {}
            for line in lines:
                text = line["text"]
                cx, cy = line["center"]

                if cy <= reject_anchor_y + 5:
                    continue

                if text in header_names:
                    vertical_distance = cy - reject_anchor_y
                    horizontal_distance = abs(cx - reject_anchor_x)
                    total_distance = (vertical_distance ** 2 + horizontal_distance ** 2) ** 0.5

                    if (
                        text not in closest_headers or
                        total_distance < closest_headers[text]["total_distance"]
                    ):
                        closest_headers[text] = {
                            "cx": cx,
                            "cy": cy,
                            "total_distance": total_distance,
                            "line": line
                        }

            for text, info in closest_headers.items():
                header_x_positions[text] = info["cx"]

        row_names = list(
            RowObjects.objects
            .filter(form_object=table)
            .values_list('row_name', flat=True)
        )

        reject_table = []

        for field in row_names:
            for line in lines:
                if (
                    field == line["text"]
                    and line["center"][0] < reject_anchor_x
                    and line["center"][1] > reject_anchor_y
                ):
                    field_y = line["center"][1]
                    label_x = line["center"][0]
                    same_row = [
                        l for l in lines
                        if abs(l["center"][1] - field_y) < 15.6 and l["center"][0] > label_x + 5
                    ]

                    field_values = {}
                    for col in header_names:
                        col_x = header_x_positions.get(col)
                        if col_x is None:
                            field_values[col] = {"text": "N/A", "confidence": 0.0}
                            continue
                        header_info = closest_headers.get(col)
                        if not header_info:
                            field_values[col] = {"text": "N/A", "confidence": 0.0}
                            continue
                        header_bbox = header_info["line"]["bbox"]
                        matches = [l for l in same_row if is_same_column(header_bbox, l["bbox"])]
                        closest = min(matches, key=lambda l: abs(l["center"][0] - col_x), default=None)
                        if closest:
                            field_values[col] = {
                                "text": closest["text"],
                                "confidence": closest.get("confidence", 1.0)
                            }
                        else:
                            field_values[col] = {"text": "N/A", "confidence": 0.0}

                    row_object = {
                        label_header: {
                            "text": field,
                            "confidence": line.get("confidence", 1.0)
                        }
                    }
                    row_object.update(field_values)
                    reject_table.append(row_object)
                    break

        tables.append({t: reject_table})
    print(tables)
    return JsonResponse({
        "extracted_table": tables,
        "extracted_fields": extra_totals
    }, json_dumps_params={"indent": 2})






def save_form(request):
    if request.method == "POST":
        data = json.loads(request.body.decode("utf-8"))
        form_template = FormName.objects.get(pk=data['template'])
        print(data)
        # --- Save to DB ---
        extraction = Extraction.objects.create(source="azure-ocr", form_name= form_template, uploaded_by=request.user)

        if data.get("extracted_fields"):
            ExtractedFields.objects.create(
                extraction=extraction,
                fields=data["extracted_fields"]
            )

        for table in data.get("extracted_table", []):
            for table_name, rows in table.items():
                ExtractedTable.objects.create(
                    extraction=extraction,
                    table_name=table_name,
                    data=rows
                )

        # --- Generate Excel ---
        wb = Workbook()
        ws = wb.active
        ws.title = "Extracted Data"

        # Styles
        header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        title_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
        bold_font = Font(bold=True)
        title_font = Font(bold=True, size=14, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Fields
        if data.get("extracted_fields"):
            ws.append(["Field", "Value"])
            for col in range(1, 3):
                cell = ws.cell(row=1, column=col)
                cell.font = bold_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border

            for field, value in data["extracted_fields"].items():
                ws.append([field, value])
                for col in range(1, 3):
                    cell = ws.cell(row=ws.max_row, column=col)
                    cell.alignment = center_align
                    cell.border = thin_border

            ws.append([])

        # Tables
        for table in data.get("extracted_table", []):
            for table_name, rows in table.items():
                if not rows:
                    continue

                # Title
                ws.append([])
                start_row = ws.max_row + 1
                num_cols = len(rows[0])
                ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=num_cols)
                title_cell = ws.cell(row=start_row, column=1, value=table_name)
                title_cell.font = title_font
                title_cell.alignment = center_align
                title_cell.fill = title_fill
                title_cell.border = thin_border

                # Headers (blank out "NA", "N/A", "na")
                headers = list(rows[0].keys())
                display_headers = [
                    "" if str(h).strip().lower() in ["na", "n/a"] else h
                    for h in headers
                ]

                ws.append(display_headers)
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=ws.max_row, column=col)
                    cell.font = bold_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                    cell.border = thin_border

                # Rows
                for row in rows:
                    row_values = [row.get(h, "") for h in headers]
                    ws.append(row_values)
                    for col in range(1, len(headers) + 1):
                        cell = ws.cell(row=ws.max_row, column=col)
                        cell.alignment = center_align
                        cell.border = thin_border

        # Auto column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2

        # Save file
        tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(tmp_file.name)

        download_url = f"/download_excel/{os.path.basename(tmp_file.name)}"
        return JsonResponse({
            "success": True,
            "download_url": download_url,
            "extraction_id": extraction.id  # ✅ return ID so you can fetch later
        })

    return JsonResponse({"success": False, "error": "Invalid request"}, status=400)


def download_excel(request, filename):
    filepath = os.path.join(tempfile.gettempdir(), filename)
    if not os.path.exists(filepath):
        return HttpResponse("File not found.", status=404)

    with open(filepath, "rb") as f:
        response = HttpResponse(
            f.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response



def get_files(request):
    return render(request, 'get_files.html')



from django.core.paginator import Paginator

@csrf_exempt
def filter_data(request):
    form_name = request.GET.get("form_name", "")
    start_date = request.GET.get("start_date", "")
    end_date = request.GET.get("end_date", "")
    page = int(request.GET.get("page", 1))
    per_page = int(request.GET.get("per_page", 6))  # cards per page

    qs = Extraction.objects.all().order_by("-created_at")

    # Filter by form_name
    if form_name:
        qs = qs.filter(form_name__name__icontains=form_name)

    # Filter by date range
    if start_date and end_date:
        try:
            start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
            end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
            qs = qs.filter(created_at__date__range=(start_date_obj, end_date_obj))
        except ValueError:
            return JsonResponse({"error": "Invalid date format. Use YYYY-MM-DD"}, status=400)

    # Pagination
    paginator = Paginator(qs, per_page)
    page_obj = paginator.get_page(page)

    results = []
    for extraction in page_obj.object_list:
       
        extraction_dict = {
            "template": extraction.form_name.name if extraction.form_name else None,
            "extracted_table": [],
            "extracted_fields": {},
            "id": extraction.pk
        }

        # Tables
        for table in extraction.tables.all():
            extraction_dict["extracted_table"].append({
                table.table_name: table.data
            })

        # Fields
        for field in extraction.fields.all():
            
            for key, value in field.fields.items():
                extraction_dict["extracted_fields"][key] = value

        results.append(extraction_dict)

    return JsonResponse({
        "results": results,
        "pagination": {
            "page": page_obj.number,
            "total_pages": paginator.num_pages,
            "has_next": page_obj.has_next(),
            "has_prev": page_obj.has_previous()
        }
    }, safe=False, json_dumps_params={"indent": 4})



@login_required
def form_detail(request, pk=None):
    try:
        extraction = Extraction.objects.get(pk=pk)
    except Extraction.DoesNotExist:
        return render(request, "404.html", status=404)

    extraction_dict = {
        "id": extraction.pk,
        "template": extraction.form_name.name if extraction.form_name else None,
        "extracted_table": [],
        "extracted_fields": {}
    }

   
    for table in extraction.tables.all():
        extraction_dict["extracted_table"].append({
            table.table_name: table.data
        })

   
    for field in extraction.fields.all():
        for key, value in field.fields.items():
            extraction_dict["extracted_fields"][key] = value

    return render(request, "form_details.html", {"extraction": extraction_dict})



@csrf_exempt
def quantity_graded_eggs(request):
    with open("main/2nd img res.json") as f:
        ocr_data = json.load(f)

    def get_center(bbox):
        xs = [p["x"] for p in bbox]
        ys = [p["y"] for p in bbox]
        return (sum(xs) / 4, sum(ys) / 4)

    headers = [
        "JUMBO", "XLARGE", "LARGE", "MEDIUM", "SMALL", "XSMALL", "PEEWEE",
        "CRACKED", "DIRTY", "SHELL ONLY", "SOFT SHELL", "SPOILED", "ASSORTED", "S-JUMBO"
    ]

    lines = []
    for block in ocr_data["readResult"]["blocks"]:
        for line in block["lines"]:
            center = get_center(line["boundingPolygon"])
            lines.append({
                "text": line["text"].strip(),
                "center": center,
                "bbox": line["boundingPolygon"]
            })

    results = {}
    y_tolerance = 10

    for line in lines:

        label = line["text"].upper()
        if label in headers:
            label_x, label_y = line["center"]

            #find closest item to the right and roughly same y
            right_candidates = [
                l for l in lines
                if l["center"][0] > label_x and abs(l["center"][1] - label_y) <= y_tolerance
            ]
            right_candidates.sort(key=lambda l: l["center"][0])  # sort by x distance

            if right_candidates:
                results[label] = right_candidates[0]["text"]
            else:
                results[label] = None

    return JsonResponse(results)





def extract_table_rows_from_file(request):
    with open("main/samp4json.json") as f:
        ocr_data = json.load(f)


    lines = []
    for block in ocr_data["readResult"]["blocks"]:
        for line in block["lines"]:
            center = get_center(line["boundingPolygon"])
            lines.append({
                "text": line["text"].strip(),
                "center": center,
                "bbox": line["boundingPolygon"]
            })


    header = next((line for line in lines if "COLLECTED GOOD EGGS" in line["text"].upper()), None)
    if not header:
        print("Header 'COLLECTED GOOD EGGS' not found.")
        return []

    _, header_y = header["center"]
    column_lines = [line for line in lines if line["center"][1] > header_y + 10]

    from collections import defaultdict
    rows_dict = defaultdict(list)
    for line in column_lines:
        y_bucket = round(line["center"][1] / 10) * 10  # bucket by y to group rows
        rows_dict[y_bucket].append(line)

    sorted_rows = []
    for y in sorted(rows_dict):
        row_items = sorted(rows_dict[y], key=lambda l: l["center"][0])  # sort by x (left to right)
        if len(row_items) >= 3:
            sorted_rows.append({
                "size": row_items[0]["text"],
                "description": row_items[1]["text"],
                "total": row_items[2]["text"]
            })


    

    return JsonResponse(sorted_rows, safe=False)

#extract_table_by_headers2
# def extract_table_by_headers2(request):
#     # table_name = "Business Name/Style ANDOKS LITSON CORPORATION Payment Terms: July 07, 2025"
#     table_name = "EPP PRODUCTION LOT-BATCHCODE REPORT"
#     # table_name = "Customer Id: $09-145"
#     # table_name = "Table 1 Title"

#     # expected_headers = ["Part Number", "Description", "UOM", "Delivered"]
#     # expected_headers = ["Part Number", "Description", "Transferred", "\"Transferred", "Received", "Returned", "Retumed"]
#     expected_headers = ["PACKER","PROD'N DATE", "HOUSE NUMBER", "MOTHER SKU", "INPUT (PCS)", "TRANSFORMATION", "TOTAL OTY"]
#     # expected_headers = ["Header 1", "Header 2", "Header 3", "Header 4", "Header 5"]
    

#     # with open("main/samp9.json") as f:
#     # with open("main/samp8b.json") as f:
#     with open("main/samp5.json") as f:
#     # with open("main/samp10.json") as f:
#         ocr_data = json.load(f)

#     lines = []
#     for block in ocr_data["readResult"]["blocks"]:
#         for line in block["lines"]:
#             center = get_center(line["boundingPolygon"])
#             lines.append({
#                 "text": line["text"].strip(),
#                 "center": center,
#                 "bbox": line["boundingPolygon"]
#             })

#     table_title = table_name.strip().upper()
#     expected_headers = [h.strip().upper() for h in expected_headers]


#     anchor_y = None
#     for line in lines:
#         if line["text"].strip().upper() == table_title:
#             _, anchor_y = line["center"]
#             break
#     if anchor_y is None:
#         return {"error": "Table title not found"}

#     headers_in_doc = {}
#     for line in lines:
#         text = line["text"].strip().upper()
#         cx, cy = line["center"]
#         if abs(cy - anchor_y) < 100 and text in expected_headers:
#             headers_in_doc[text] = cx
#             print(cy)
#     print(headers_in_doc)
#     if len(headers_in_doc) < len(expected_headers):

#         return {"error": "Some headers not found below title"}


#     extracted_rows = []
#     row_lines = [line for line in lines if line["center"][1] > anchor_y + 50]

#     used_rows = set()
    
#     stop_text = "BAWAL MAG IWAN NG CRATES"

#     for line in row_lines:
#         row_y = line["center"][1]

#         # STOP if the special line is detected
#         if line["text"].strip().upper() == stop_text.upper():
#             break

#         if any(abs(row_y - y) < 20 for y in used_rows):
#             continue

#         same_row = [l for l in row_lines if abs(l["center"][1] - row_y) < 10]

#         # Also stop if the special text is found anywhere in this row
#         if any(stop_text.upper() in l["text"].strip().upper() for l in same_row):
#             break

#         used_rows.add(row_y)

#         row_data = {}
#         for header in expected_headers:
#             col_x = headers_in_doc.get(header)
#             if col_x is None:
#                 row_data[header] = "N/A"
#                 continue

#             header_bbox = next(line["bbox"] for line in lines if line["text"].strip().upper() == header)
#             matches = [l for l in same_row if is_same_column(header_bbox, l["bbox"])]
#             closest = min(matches, key=lambda l: abs(l["center"][0] - col_x), default=None)
#             row_data[header] = closest["text"] if closest else "N/A"

#         extracted_rows.append(row_data)


#     return JsonResponse(extracted_rows, safe=False)



def extract_table_by_headers2(request):
    # table_name = "Customer Id: $09-111"
    table_name = "Customer Id: $09-145"
    # table_name = "Business Name/Style ANDOKS LITSON CORPORATION Payment Terms: July 07, 2025"
    # table_name = "EPP PRODUCTION LOT-BATCHCODE REPORT"
    expected_headers = ["Part Number", "Description", "Transferred", "\"Transferred", "Received", "Returned", "Retumed"]
    # expected_headers = ["Part Number", "Description", "UOM", "Delivered"]
    # expected_headers = ["PACKER","PROD'N DATE", "HOUSE NUMBER", "MOTHER SKU", "INPUT (PCS)", "TRANSFORMATION", "-", "SALLE ORDER", "UOM","TOTAL OTY","DIRTY (PCS)",]

    with open("main/samp9.json") as f:
    # with open("main/samp8b.json") as f:
    # with open("main/samp5.json") as f:
    # with open("main/samp10.json") as f:
    # with open("main/samp11.json") as f:
        ocr_data = json.load(f)

    lines = []
    for block in ocr_data["readResult"]["blocks"]:
        for line in block["lines"]:
            center = get_center(line["boundingPolygon"])
            lines.append({
                "text": line["text"].strip(),
                "center": center,
                "bbox": line["boundingPolygon"]
            })

    table_title = table_name.strip().upper()
    expected_headers_upper = [h.strip().upper() for h in expected_headers]

    # Find table title Y
    anchor_y = None
    for line in lines:
        if line["text"].strip().upper() == table_title:
            _, anchor_y = line["center"]
            break
    if anchor_y is None:
        return {"error": "Table title not found"}

    # Find headers in the document
    headers_in_doc = {}
    header_y_positions = []
    for line in lines:
        text = line["text"].strip().upper()
        cx, cy = line["center"]
        if cy > anchor_y and text in expected_headers_upper:
            headers_in_doc[text] = cx
            header_y_positions.append(cy)

    # Warn if some headers are missing but do not stop execution
    found_headers = list(headers_in_doc.keys())
    missing_headers = [h for h in expected_headers_upper if h not in found_headers]
    if missing_headers:
        print(f"⚠ Missing headers: {missing_headers}")
    if not headers_in_doc:
        return {"error": "No matching headers found below title"}

    # Dynamic start position
    if header_y_positions:
        start_y = max(header_y_positions) + 5
    else:
        start_y = anchor_y + 50

    extracted_rows = []
    row_lines = [line for line in lines if line["center"][1] > start_y]

    used_rows = set()
    stop_text = "BAWAL MAG IWAN NG CRATES"

    for line in row_lines:
        row_y = line["center"][1]

        # STOP if the special line is detected
        if line["text"].strip().upper() == stop_text.upper():
            break

        if any(abs(row_y - y) < 20 for y in used_rows):
            continue

        same_row = [l for l in row_lines if abs(l["center"][1] - row_y) < 13.5]

        # Also stop if the special text is found anywhere in this row
        if any(stop_text.upper() in l["text"].strip().upper() for l in same_row):
            break

        used_rows.add(row_y)

        row_data = {}
        for header_text_upper, col_x in headers_in_doc.items():
            # Use the original header casing from expected_headers
            original_header = next((h for h in expected_headers if h.strip().upper() == header_text_upper), header_text_upper)
            header_bbox = next(line["bbox"] for line in lines if line["text"].strip().upper() == header_text_upper)
            matches = [l for l in same_row if is_same_column(header_bbox, l["bbox"])]
            closest = min(matches, key=lambda l: abs(l["center"][0] - col_x), default=None)
            row_data[original_header] = closest["text"] if closest else "N/A"

        extracted_rows.append(row_data)


    return extracted_rows


